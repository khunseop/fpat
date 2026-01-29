import os
import time
import datetime
import logging
import requests
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# SSL 설정 (urllib3 버전 호환성 고려)
try:
    # 이전 버전의 urllib3에서 사용
    requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS += ':DES-CBC3-SHA'
except AttributeError:
    # 최신 버전의 urllib3에서는 해당 속성이 없음
    pass
requests.packages.urllib3.disable_warnings()

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


def apply_excel_style(file_name: str) -> None:
    """
    모든 시트에 대해 헤더에 연한 회색 배경을 적용하고,
    헤더의 너비를 자동 조절하되 최대 너비를 40으로 제한합니다.
    
    :param file_name: 처리할 엑셀 파일 이름
    """
    try:
        workbook = load_workbook(file_name)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # 모든 워크시트를 순회하며 스타일 적용
        for worksheet in workbook.worksheets:
            header_row = worksheet[1]
            for cell in header_row:
                cell.fill = header_fill
                column_letter = cell.column_letter
                max_length = 0
                for col_cell in worksheet[column_letter]:
                    if col_cell.value is not None:
                        cell_length = len(str(col_cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(40, adjusted_width)

        workbook.save(file_name)
    except Exception as error:
        logging.error("엑셀 스타일 적용 중 오류 발생: %s", error)


class PaloAltoAPI:
    def __init__(self, hostname: str, username: str, password: str) -> None:
        self.hostname = hostname
        self.base_url = f'https://{hostname}/api/'
        self.api_key = self._get_api_key(username, password)

    def save_to_excel(self, data, sheet_names=None) -> str:
        """
        단일 DataFrame 또는 DataFrame 리스트를 엑셀 파일로 저장합니다.
        파일 이름은 현재 날짜, 호스트명, 시트명을 활용하여 자동 생성됩니다.
        
        :param data: 저장할 DataFrame 또는 DataFrame 리스트
        :param sheet_names: 단일 시트명(str) 또는 시트명 리스트 (옵션)
                            리스트가 아닌 경우 단일 시트로 저장됩니다.
                            기본값은 단일 시트의 경우 "Sheet1",
                            다중 시트의 경우 "Sheet1", "Sheet2", ... 로 지정됩니다.
        :return: 생성된 엑셀 파일 이름
        """
        current_date = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        
        # 단일 DataFrame인 경우
        if not isinstance(data, list):
            sheet_name = sheet_names if isinstance(sheet_names, str) else "Sheet1"
            file_name = f"{current_date}_{self.hostname}_{sheet_name}.xlsx"
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # 여러 DataFrame인 경우
            num_sheets = len(data)
            if sheet_names is None:
                sheet_names = [f"Sheet{i+1}" for i in range(num_sheets)]
            elif not isinstance(sheet_names, list):
                sheet_names = [sheet_names]
            file_name = f"{current_date}_{self.hostname}_combined.xlsx"
            with pd.ExcelWriter(file_name) as writer:
                for df, sheet in zip(data, sheet_names):
                    df.to_excel(writer, sheet_name=sheet, index=False)
        
        # 엑셀 파일에 스타일 적용 (모든 시트)
        apply_excel_style(file_name)
        return file_name

    @staticmethod
    def _get_member_texts(xml_elements) -> list:
        """
        주어진 XML 요소 리스트에서 텍스트 값을 추출합니다.

        :param xml_elements: XML 요소 리스트
        :return: 텍스트 값 리스트
        """
        try:
            return [element.text for element in xml_elements if element.text is not None]
        except Exception:
            return []

    @staticmethod
    def list_to_string(list_data: list) -> str:
        """
        리스트 데이터를 콤마 구분 문자열로 변환합니다.

        :param list_data: 리스트 데이터
        :return: 콤마로 구분된 문자열
        """
        return ','.join(str(item) for item in list_data)

    def get_api_data(self, parameters, timeout: int = 10000):
        """API 호출을 수행합니다."""
        try:
            response = requests.get(
                self.base_url,
                params=parameters,
                verify=False,
                timeout=timeout
            )
            if response.status_code != 200:
                raise Exception(f"API 요청 실패 (상태 코드: {response.status_code}): {response.text}")
            return response
        except requests.exceptions.Timeout:
            raise Exception("API 요청 시간 초과")
        except requests.exceptions.ConnectionError:
            raise Exception(f"API 서버 연결 실패: {self.hostname}")
        except requests.exceptions.RequestException as e:
            raise Exception(f"API 요청 중 오류 발생: {str(e)}")

    def _get_api_key(self, username: str, password: str) -> str:
        """API 키를 생성합니다."""
        try:
            keygen_params = (
                ('type', 'keygen'),
                ('user', username),
                ('password', password)
            )
            response = self.get_api_data(keygen_params)
            tree = ET.fromstring(response.text)
            key_element = tree.find('./result/key')
            if key_element is None:
                raise Exception("API 키를 찾을 수 없습니다")
            return key_element.text
        except ET.ParseError:
            raise Exception("API 응답 XML 파싱 실패")
        except Exception as e:
            raise Exception(f"API 키 생성 실패: {str(e)}")

    def get_vsys_list(self) -> list:
        """
        vsys 리스트를 반환합니다.

        :return: vsys 이름 리스트
        """
        params = (
            ('key', self.api_key),
            ('type', 'config'),
            ('action', 'get'),
            ('xpath', '/config/devices/entry/vsys/entry'),
        )

        response = self.get_api_data(params)
        vsys_entries = ET.fromstring(response.text).findall('./result/entry')
        return [vsys.attrib.get('name') for vsys in vsys_entries]

    def get_config(self, config_type: str = 'running') -> str:
        """
        설정 정보를 가져옵니다.

        :param config_type: 'running' 또는 기타
        :return: 설정 XML 문자열
        """
        action = 'show' if config_type == 'running' else 'get'
        params = (
            ('key', self.api_key),
            ('type', 'config'),
            ('action', action),
            ('xpath', '/config')
        )
        response = self.get_api_data(params)
        return response.text

    def save_config(self, config_type: str = 'running') -> bool:
        """
        설정 정보를 XML 파일로 저장합니다.

        :param config_type: 'running' 또는 기타
        :return: 저장 성공 여부
        """
        current_date = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        config_data = self.get_config(config_type)
        file_name = f'{current_date}_{self.hostname}_{config_type}_config.xml'
        try:
            with open(file_name, mode='w', encoding='utf8') as file:
                file.write(config_data)
            return True
        except Exception as error:
            logging.error("설정 저장 중 오류 발생: %s", error)
            return False

    def get_system_info(self) -> pd.DataFrame:
        """
        시스템 정보를 DataFrame으로 반환합니다.

        :return: 시스템 정보 DataFrame
        """
        params = (
            ('type', 'op'),
            ('cmd', '<show><system><info/></system></show>'),
            ('key', self.api_key)
        )
        response = self.get_api_data(params)
        tree = ET.fromstring(response.text)
        uptime = tree.findtext("./result/system/uptime")
        info = {
            "hostname": tree.findtext("./result/system/hostname"),
            "ip_address": tree.findtext("./result/system/ip-address"),
            "mac_address": tree.findtext("./result/system/mac-address"),
            "uptime": uptime.split(" ")[0] if uptime else None,
            "model": tree.findtext("./result/system/model"),
            "serial_number": tree.findtext("./result/system/serial"),
            "sw_version": tree.findtext("./result/system/sw-version"),
            "app_version": tree.findtext("./result/system/app-version"),
        }
        return pd.DataFrame(info, index=[0])

    def get_system_state(self) -> pd.DataFrame:
        """
        시스템 상태 정보를 DataFrame으로 반환합니다.

        :return: 시스템 상태 DataFrame
        """
        params = (
            ('type', 'op'),
            ('cmd', '<show><system><state><filter>cfg.general.max*</filter></state></system></show>'),
            ('key', self.api_key)
        )
        response = self.get_api_data(params)
        tree = ET.fromstring(response.text)
        result_text = tree.findtext("./result")

        # 기본값 초기화
        max_address = max_address_group = max_service = max_service_group = max_policy_rule = ""
        if result_text:
            for line in result_text.split('\n'):
                if line.startswith('cfg.general.max-address:'):
                    max_address = line.split(': ')[1]
                elif line.startswith('cfg.general.max-address-group:'):
                    max_address_group = line.split(': ')[1]
                elif line.startswith('cfg.general.max-service:'):
                    max_service = line.split(': ')[1]
                elif line.startswith('cfg.general.max-service-group:'):
                    max_service_group = line.split(': ')[1]
                elif line.startswith('cfg.general.max-policy-rule:'):
                    max_policy_rule = line.split(': ')[1]

        state = {
            "hostname": self.hostname,
            "max_policy_rule": max_policy_rule,
            "max_address": max_address,
            "max_address_group": max_address_group,
            "max_service": max_service,
            "max_service_group": max_service_group
        }
        return pd.DataFrame(state, index=[0])

    def export_security_rules(self, config_type: str = 'running') -> pd.DataFrame:
        """
        보안 규칙 정보를 DataFrame으로 반환합니다.

        :param config_type: 'running' 또는 기타
        :return: 보안 규칙 DataFrame
        """
        config_xml = self.get_config(config_type)
        tree = ET.fromstring(config_xml)
        vsys_entries = tree.findall('./result/config/devices/entry/vsys/entry')
        security_rules = []

        for vsys in vsys_entries:
            vsys_name = vsys.attrib.get('name')
            rulebase = vsys.findall('./rulebase/security/rules/entry')
            for idx, rule in enumerate(rulebase):
                rule_name = str(rule.attrib.get('name'))
                disabled_list = self._get_member_texts(rule.findall('./disabled'))
                disabled_status = "N" if self.list_to_string(disabled_list) == "yes" else "Y"
                action = self.list_to_string(self._get_member_texts(rule.findall('./action')))
                source = self.list_to_string(self._get_member_texts(rule.findall('./source/member')))
                user = self.list_to_string(self._get_member_texts(rule.findall('./source-user/member')))
                destination = self.list_to_string(self._get_member_texts(rule.findall('./destination/member')))
                service = self.list_to_string(self._get_member_texts(rule.findall('./service/member')))
                application = self.list_to_string(self._get_member_texts(rule.findall('./application/member')))
                url_filtering = self.list_to_string(self._get_member_texts(rule.findall('./profile-setting/profiles/url-filtering/member')))
                category = self.list_to_string(self._get_member_texts(rule.findall('./category/member')))
                category = "any" if not category else category
                description_list = self._get_member_texts(rule.findall('./description'))
                description = self.list_to_string([desc.replace('\n', ' ') for desc in description_list])

                rule_info = {
                    "Vsys": vsys_name,
                    "Seq": idx + 1,
                    "Rule Name": rule_name,
                    "Enable": disabled_status,
                    "Action": action,
                    "Source": source,
                    "User": user,
                    "Destination": destination,
                    "Service": service,
                    "Application": application,
                    "Security Profile": url_filtering,
                    "Category": category,
                    "Description": description,
                }
                security_rules.append(rule_info)

        return pd.DataFrame(security_rules)

    def export_network_objects(self, config_type: str = 'running') -> pd.DataFrame:
        """
        네트워크 객체 정보를 DataFrame으로 반환합니다.

        :param config_type: 'running' 또는 기타
        :return: 네트워크 객체 DataFrame
        """
        config_xml = self.get_config(config_type)
        tree = ET.fromstring(config_xml)
        address_entries = tree.findall('./result/config/devices/entry/vsys/entry/address/entry')
        address_objects = []

        for address in address_entries:
            address_name = address.attrib.get('name')
            address_type = address.find('*').tag if address.find('*') is not None else ""
            member_elements = address.findall(f'./{address_type}')
            members = [elem.text for elem in member_elements if elem.text is not None]

            object_info = {
                "Name": address_name,
                "Type": address_type,
                "Value": self.list_to_string(members)
            }
            address_objects.append(object_info)

        return pd.DataFrame(address_objects)

    def export_network_group_objects(self, config_type: str = 'running') -> pd.DataFrame:
        """
        네트워크 그룹 객체 정보를 DataFrame으로 반환합니다.

        :param config_type: 'running' 또는 기타
        :return: 네트워크 그룹 객체 DataFrame
        """
        config_xml = self.get_config(config_type)
        tree = ET.fromstring(config_xml)
        group_entries = tree.findall('./result/config/devices/entry/vsys/entry/address-group/entry')
        group_objects = []

        for group in group_entries:
            group_name = group.attrib.get('name')
            member_elements = group.findall('./static/member')
            members = [elem.text for elem in member_elements if elem.text is not None]

            group_info = {
                "Group Name": group_name,
                "Entry": self.list_to_string(members)
            }
            group_objects.append(group_info)

        return pd.DataFrame(group_objects)

    def export_service_objects(self, config_type: str = 'running') -> pd.DataFrame:
        """
        서비스 객체 정보를 DataFrame으로 반환합니다.

        :param config_type: 'running' 또는 기타
        :return: 서비스 객체 DataFrame
        """
        config_xml = self.get_config(config_type)
        tree = ET.fromstring(config_xml)
        service_entries = tree.findall('./result/config/devices/entry/vsys/entry/service/entry')
        service_objects = []

        for service in service_entries:
            service_name = service.attrib.get('name')
            protocol_elem = service.find('protocol')
            if protocol_elem is not None:
                for protocol in protocol_elem:
                    protocol_name = protocol.tag
                    port = protocol.find('port').text if protocol.find('port') is not None else None

                    service_info = {
                        "Name": service_name,
                        "Protocol": protocol_name,
                        "Port": port,
                    }
                    service_objects.append(service_info)

        return pd.DataFrame(service_objects)

    def export_service_group_objects(self, config_type: str = 'running') -> pd.DataFrame:
        """
        서비스 그룹 객체 정보를 DataFrame으로 반환합니다.

        :param config_type: 'running' 또는 기타
        :return: 서비스 그룹 객체 DataFrame
        """
        config_xml = self.get_config(config_type)
        tree = ET.fromstring(config_xml)
        group_entries = tree.findall('./result/config/devices/entry/vsys/entry/service-group/entry')
        group_objects = []

        for group in group_entries:
            group_name = group.attrib.get('name')
            member_elements = group.findall('./members/member')
            members = [elem.text for elem in member_elements if elem.text is not None]

            group_info = {
                "Group Name": group_name,
                "Entry": self.list_to_string(members),
            }
            group_objects.append(group_info)

        return pd.DataFrame(group_objects)

    def export_hit_count(self, vsys_name: str = 'vsys1') -> pd.DataFrame:
        """
        히트 카운트 정보를 DataFrame으로 반환합니다.

        :param vsys_name: vsys 이름
        :return: 히트 카운트 DataFrame
        """
        params = (
            ('type', 'op'),
            (
                'cmd',
                f"<show><rule-hit-count><vsys><vsys-name><entry name='{vsys_name}'>"
                "<rule-base><entry name='security'><rules><all/></rules></entry></rule-base>"
                "</entry></vsys-name></vsys></rule-hit-count></show>"
            ),
            ('key', self.api_key)
        )
        response = self.get_api_data(params)
        tree = ET.fromstring(response.text)
        rule_entries = tree.findall('./result/rule-hit-count/vsys/entry/rule-base/entry/rules/entry')

        hit_counts = []
        for rule in rule_entries:
            rule_name = str(rule.attrib.get('name'))
            member_texts = self._get_member_texts(rule)
            try:
                hit_count = member_texts[1]
                last_hit_ts = int(member_texts[2])
                first_hit_ts = int(member_texts[4])
            except (IndexError, ValueError) as error:
                logging.error("히트 카운트 파싱 중 오류 발생: %s", error)
                continue

            no_unused_days = 99999
            no_hit_date = datetime.datetime(1900, 1, 1).strftime('%Y-%m-%d')

            if first_hit_ts == 0:
                unused_days = no_unused_days
            else:
                unused_days = (datetime.datetime.now() - datetime.datetime.fromtimestamp(last_hit_ts)).days

            last_hit_date = no_hit_date if last_hit_ts == 0 else datetime.datetime.fromtimestamp(last_hit_ts).strftime('%Y-%m-%d')
            first_hit_date = no_hit_date if first_hit_ts == 0 else datetime.datetime.fromtimestamp(first_hit_ts).strftime('%Y-%m-%d')

            hit_counts.append({
                "Vsys": vsys_name,
                "Rule Name": rule_name,
                "Hit Count": hit_count,
                "First Hit Date": first_hit_date,
                "Last Hit Date": last_hit_date,
                "Unused Days": unused_days
            })

        return pd.DataFrame(hit_counts)

    def show_config_running_match_rematch(self):
        config_xml = self.get_config('running')
        tree = ET.fromstring(config_xml)

        result_text = tree.find("./result/config/devices/entry/deviceconfig/setting/config/rematch").text

        if result_text:
            return [f"rematch: {result_text}"]
        else:
            raise ValueError(f"결과값 없음")
    
    def run_command(self, command):
        try:
            params = (
                ('type', 'op'),
                ('cmd', command),
                ('key', self.api_key)
            )
            response = self.get_api_data(params)
            tree = ET.fromstring(response.text)

            member_text = tree.findtext("./result/member")
            if member_text:
                return member_text.strip().splitlines()
            
            result_node = tree.find("./result")
            if result_node is None:
                return []
            
            children = list(result_node)
            if children:
                return [
                    f"{child.tag}: {(child.text or '').strip()}"
                    for child in children
                ]
            
            result_text = (result_node.text or '').strip()
            return result_text.splitlines() if result_text else []
        except Exception as e:
            logging.error(f"명령어 실행 중 오류 발생: {e}")
            return []


    def export_nat_rules(self, config_type: str = 'running') -> pd.DataFrame:
        """
        NAT 정책 정보를 DataFrame으로 반환합니다.

        :param config_type: 'running' 또는 기타
        :return: NAT 정책 DataFrame
        """
        config_xml = self.get_config(config_type)
        tree = ET.fromstring(config_xml)
        vsys_entries = tree.findall('./result/config/devices/entry/vsys/entry')
        nat_rules = []

        for vsys in vsys_entries:
            vsys_name = vsys.attrib.get('name')
            rulebase = vsys.findall('./rulebase/nat/rules/entry')
            for idx, rule in enumerate(rulebase):
                rule_name = str(rule.attrib.get('name'))
                disabled_list = self._get_member_texts(rule.findall('./disabled'))
                disabled_status = "N" if self.list_to_string(disabled_list) == "yes" else "Y"
                nat_type = self.list_to_string(self._get_member_texts(rule.findall('./nat-type')))
                
                source_zone = self.list_to_string(self._get_member_texts(rule.findall('./from/member')))
                destination_zone = self.list_to_string(self._get_member_texts(rule.findall('./to/member')))
                destination_interface = self.list_to_string(self._get_member_texts(rule.findall('./to-interface')))
                source = self.list_to_string(self._get_member_texts(rule.findall('./source/member')))
                destination = self.list_to_string(self._get_member_texts(rule.findall('./destination/member')))
                service = self.list_to_string(self._get_member_texts(rule.findall('./service')))
                source_translation = self.list_to_string(self._get_member_texts(rule.findall('./source-translation/static-ip/translated-address')))
                destination_translation = self.list_to_string(self._get_member_texts(rule.findall('./destination-translation/translated-address')))

                description_list = self._get_member_texts(rule.findall('./description'))
                description = self.list_to_string([desc.replace('\n', ' ') for desc in description_list])

                rule_info = {
                    "Vsys": vsys_name,
                    "Seq": idx + 1,
                    "Rule Name": rule_name,
                    "Enable": disabled_status,
                    "NAT Type": nat_type,
                    "Original Packet Source Zone": source_zone,
                    "Original Packet Destination Zone": destination_zone,
                    "Original Packet Destination Interface": destination_interface,
                    "Original Packet Source Address": source,
                    "Original Packet Destination Address": destination,
                    "Service": service,
                    "Translated Packet Source Translation": source_translation,
                    "Translated Packet Destination Translation": destination_translation,
                    "Description": description,
                }
                nat_rules.append(rule_info)

        return pd.DataFrame(nat_rules)

    def get_interface_management_info(self) -> pd.DataFrame:
        """
        관리 인터페이스 정보를 DataFrame으로 반환합니다.

        :return: 관리 인터페이스 정보 DataFrame
        """
        try:
            params = (
                ('type', 'op'),
                ('cmd', '<show><interface>management</interface></show>'),
                ('key', self.api_key)
            )
            response = self.get_api_data(params)
            tree = ET.fromstring(response.text)
            
            info_elem = tree.find("./result/info")
            if info_elem is None:
                logging.warning("관리 인터페이스 정보를 찾을 수 없습니다")
                return pd.DataFrame()
            
            interface_info = {
                "hostname": self.hostname,
                "name": info_elem.findtext("./name") or "",
                "state": info_elem.findtext("./state") or "",
                "ip": info_elem.findtext("./ip") or "",
                "netmask": info_elem.findtext("./netmask") or "",
                "speed": info_elem.findtext("./speed") or "",
                "duplex": info_elem.findtext("./duplex") or "",
                "link_state": info_elem.findtext("./link-state") or "",
            }
            
            return pd.DataFrame(interface_info, index=[0])
        except Exception as error:
            logging.error("관리 인터페이스 정보 조회 중 오류 발생: %s", error)
            return pd.DataFrame()

    def get_resource_usage(self, days: int = 1) -> pd.DataFrame:
        """
        리소스 사용률 정보를 DataFrame으로 반환합니다.

        :param days: 조회할 일수 (기본값: 1)
        :return: 리소스 사용률 DataFrame
        """
        try:
            params = (
                ('type', 'op'),
                ('cmd', f'<show><running><resource-monitor><day><last>{days}</last></day></resource-monitor></running></show>'),
                ('key', self.api_key)
            )
            response = self.get_api_data(params)
            tree = ET.fromstring(response.text)
            
            resource_data = []
            result_elem = tree.find("./result")
            if result_elem is None:
                logging.warning("리소스 사용률 정보를 찾을 수 없습니다")
                return pd.DataFrame()
            
            # CPU 사용률 정보 추출
            cpu_elem = result_elem.find("./cpu")
            if cpu_elem is not None:
                cpu_info = {
                    "hostname": self.hostname,
                    "resource_type": "CPU",
                    "average": cpu_elem.findtext("./average") or "",
                    "maximum": cpu_elem.findtext("./maximum") or "",
                    "minimum": cpu_elem.findtext("./minimum") or "",
                }
                resource_data.append(cpu_info)
            
            # 메모리 사용률 정보 추출
            memory_elem = result_elem.find("./memory")
            if memory_elem is not None:
                memory_info = {
                    "hostname": self.hostname,
                    "resource_type": "Memory",
                    "average": memory_elem.findtext("./average") or "",
                    "maximum": memory_elem.findtext("./maximum") or "",
                    "minimum": memory_elem.findtext("./minimum") or "",
                }
                resource_data.append(memory_info)
            
            # 세션 사용률 정보 추출
            session_elem = result_elem.find("./session")
            if session_elem is not None:
                session_info = {
                    "hostname": self.hostname,
                    "resource_type": "Session",
                    "average": session_elem.findtext("./average") or "",
                    "maximum": session_elem.findtext("./maximum") or "",
                    "minimum": session_elem.findtext("./minimum") or "",
                }
                resource_data.append(session_info)
            
            if not resource_data:
                logging.warning("리소스 사용률 데이터가 없습니다")
                return pd.DataFrame()
            
            return pd.DataFrame(resource_data)
        except Exception as error:
            logging.error("리소스 사용률 조회 중 오류 발생: %s", error)
            return pd.DataFrame()

    def get_system_state_detailed(self) -> pd.DataFrame:
        """
        상세 시스템 상태 정보를 DataFrame으로 반환합니다.

        :return: 시스템 상태 정보 DataFrame
        """
        try:
            params = (
                ('type', 'op'),
                ('cmd', '<show><system><state/></system></show>'),
                ('key', self.api_key)
            )
            response = self.get_api_data(params)
            tree = ET.fromstring(response.text)
            
            result_elem = tree.find("./result")
            if result_elem is None:
                logging.warning("시스템 상태 정보를 찾을 수 없습니다")
                return pd.DataFrame()
            
            # 시스템 상태 정보를 딕셔너리로 수집
            state_info = {"hostname": self.hostname}
            
            # 주요 설정 제한값 추출
            result_text = result_elem.text
            if result_text:
                for line in result_text.split('\n'):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # cfg.general.max-* 패턴 추출
                    if ':' in line:
                        key, value = line.split(':', 1)
                        key = key.strip().replace('cfg.general.', '').replace('-', '_')
                        state_info[key] = value.strip()
            
            # XML 요소에서 직접 추출 가능한 정보
            for child in result_elem:
                tag_name = child.tag.replace('-', '_')
                text_value = child.text.strip() if child.text else ""
                if text_value:
                    state_info[tag_name] = text_value
            
            if len(state_info) == 1:  # hostname만 있는 경우
                logging.warning("시스템 상태 데이터가 없습니다")
                return pd.DataFrame()
            
            return pd.DataFrame(state_info, index=[0])
        except Exception as error:
            logging.error("시스템 상태 조회 중 오류 발생: %s", error)
            return pd.DataFrame()