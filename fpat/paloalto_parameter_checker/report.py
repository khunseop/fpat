#!/usr/bin/env python3
"""
Excel 리포트 생성 모듈
"""

import os
import json
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ReportGenerator:
    def __init__(self, reports_dir: str = "reports"):
        self.reports_dir = reports_dir
        os.makedirs(reports_dir, exist_ok=True)
    
    def generate_excel_report(self, results: List[Dict], summary: Dict, 
                             filename: str = None) -> Dict:
        """Excel 리포트 생성"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"palo_alto_check_report_{timestamp}.xlsx"
            
            filepath = os.path.join(self.reports_dir, filename)
            
            # 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Parameter Check Report"
            
            # 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 제목 및 요약 정보
            ws['A1'] = "Palo Alto Parameter Check Report"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            ws['A3'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"총 매개변수: {summary['total']}"
            ws['A5'] = f"정상: {summary['pass']}"
            ws['A6'] = f"실패: {summary['fail']}"
            ws['A7'] = f"오류: {summary['error']}"
            
            # 헤더 행
            headers = ['파라미터', '기대값', '현재값', '상태', '조회 방법', '변경 방법']
            start_row = 9
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # 데이터 행
            for row_idx, result in enumerate(results, start_row + 1):
                # 데이터 입력
                ws.cell(row=row_idx, column=1, value=result['parameter'])
                ws.cell(row=row_idx, column=2, value=result['expected'])
                ws.cell(row=row_idx, column=3, value=result['current'])
                ws.cell(row=row_idx, column=4, value=result['status'])
                ws.cell(row=row_idx, column=5, value=result['query_method'])
                ws.cell(row=row_idx, column=6, value=result['modify_method'])
                
                # 상태에 따른 색상 적용
                status_fill = None
                if result['status'] == 'PASS':
                    status_fill = pass_fill
                elif result['status'] == 'FAIL':
                    status_fill = fail_fill
                elif result['status'] == 'ERROR':
                    status_fill = error_fill
                
                # 행 전체에 스타일 적용
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    if status_fill:
                        cell.fill = status_fill
            
            # 열 너비 자동 조정
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except Exception as e:
                        import logging
                        logging.debug(f"Row value parsing skipped: {e}")
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(filepath)
            
            return {
                'success': True,
                'message': 'Excel 리포트 생성 완료',
                'filename': filename,
                'filepath': filepath
            }
            
        except Exception as e:
            import logging
            logging.error(f"Excel 리포트 생성 중 예외 발생: {str(e)}")
            return {
                'success': False,
                'message': f'Excel 리포트 생성 실패: {str(e)}'
            }
    

    
    def cleanup_old_reports(self, days_old: int = 1):
        """오래된 리포트 파일 정리"""
        try:
            import time
            import logging
            
            current_time = time.time()
            cutoff_time = current_time - (days_old * 24 * 60 * 60)
            
            for filename in os.listdir(self.reports_dir):
                filepath = os.path.join(self.reports_dir, filename)
                if os.path.isfile(filepath):
                    file_time = os.path.getmtime(filepath)
                    if file_time < cutoff_time:
                        os.remove(filepath)
                        
        except Exception as e:
            import logging
            logging.warning(f"오래된 리포트 정리 중 오류 발생: {str(e)}")
            pass  # 정리 실패는 무시